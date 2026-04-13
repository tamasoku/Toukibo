import re

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="玉川測量設計 登記簿データ抽出・現場データ生成", layout="wide")
st.title("登記簿データ抽出・現場データ生成ツール")
st.markdown("CSVファイルをアップロードし、所有者抽出・JUSTDB用Excel出力・現場データ生成までを一括で行えます。")

# ── Step 1: CSV読み込み・現所有者抽出 ──
st.header("Step 1: CSV読み込み・現所有者抽出")
uploaded_file = st.file_uploader("CSVファイルをアップロードしてください", type="csv")

if uploaded_file is None:
    st.stop()

# CSVファイルを読み込み (Shift-JIS / UTF-8 自動判定)
try:
    df = pd.read_csv(uploaded_file, encoding='shift_jis', dtype={'不動産番号': str})
except UnicodeDecodeError:
    uploaded_file.seek(0)
    try:
        df = pd.read_csv(uploaded_file, encoding='utf-8', dtype={'不動産番号': str})
    except Exception:
        st.error("ファイルの読み込みに失敗しました。エンコードが正しくありません。")
        st.stop()

# 必要なカラムの存在チェック
required_cols = ['権利部（甲区）氏名']
missing = [c for c in required_cols if c not in df.columns]
if missing:
    st.error(f"必要なカラムがありません: {', '.join(missing)}")
    st.stop()

# 物件特定カラムのみ前方補完（権利部カラムはffillしない）
property_cols = ['不動産番号', '所在', '地番', '地目', '地積']
property_cols = [c for c in property_cols if c in df.columns]
df_filled = df.copy()
df_filled[property_cols] = df[property_cols].ffill()

if '不動産番号' in df_filled.columns:
    df_filled['不動産番号'] = df_filled['不動産番号'].apply(lambda x: str(x))


def extract_current_owners(df_src):
    has_chiban_col = '地番' in df_src.columns
    has_purpose_col = '権利部（甲区）登記の目的' in df_src.columns
    has_junni_col = '権利部（甲区）順位番号' in df_src.columns
    has_cause_col = '権利部（甲区）原因' in df_src.columns
    has_fudosan_col = '不動産番号' in df_src.columns
    has_shozai_col = '所在' in df_src.columns

    if not has_chiban_col:
        return df_src.dropna(subset=['権利部（甲区）氏名'])

    # 不動産番号があれば最優先（全国一意のID）。無ければ所在+地番でグループ化。
    if has_fudosan_col:
        group_key = '不動産番号'
    elif has_shozai_col:
        group_key = ['所在', '地番']
    else:
        group_key = '地番'

    results = []
    for _, group in df_src.groupby(group_key, sort=False):
        if has_purpose_col:
            purposes = group['権利部（甲区）登記の目的'].dropna().unique()
            if len(purposes) == 1 and purposes[0] == '所有権敷地権':
                row = group.iloc[0].copy()
                row['権利部（甲区）氏名'] = '所有権敷地権'
                results.append(pd.DataFrame([row]))
                continue

        named = group.dropna(subset=['権利部（甲区）氏名'])
        if named.empty:
            continue

        if has_purpose_col and has_junni_col:
            work = named.copy()
            work['権利部（甲区）順位番号'] = work['権利部（甲区）順位番号'].ffill()
            if has_cause_col:
                work['権利部（甲区）原因'] = work['権利部（甲区）原因'].ffill()
            current_rows = {}
            for _, entry_group in work.groupby('権利部（甲区）順位番号', sort=False):
                purpose = str(entry_group.iloc[0].get('権利部（甲区）登記の目的', ''))
                new_rows = {}
                for idx, row in entry_group.iterrows():
                    new_rows[row['権利部（甲区）氏名']] = row

                if purpose in ('所有権移転', '所有権登記', '合併による所有権登記') or '共有者全員持分全部移転' in purpose:
                    current_rows = new_rows.copy()
                else:
                    matches = re.findall(r'([^、]+?)持分(全部|一部)', purpose)
                    if matches and '移転' in purpose:
                        for person, transfer_type in matches:
                            if transfer_type == '全部':
                                current_rows.pop(person, None)
                        current_rows.update(new_rows)
                    else:
                        current_rows.update(new_rows)

            if current_rows:
                results.append(pd.DataFrame(list(current_rows.values())))
            continue

        if has_cause_col:
            has_cause = named[named['権利部（甲区）原因'].notna()]
            if not has_cause.empty:
                last_cause_idx = has_cause.index[-1]
                results.append(named.loc[named.index >= last_cause_idx])
                continue

        results.append(named.tail(1))

    if results:
        return pd.concat(results)
    return pd.DataFrame(columns=df_src.columns)


df_current = extract_current_owners(df_filled)

with st.expander(f"アップロードデータプレビュー（全{len(df_filled)}件）"):
    st.dataframe(df_filled)

st.success(f"現所有者 {len(df_current)} 件を抽出しました。")

# ── Step 2: 名前フィルタ・JUSTDB用Excel出力 ──
st.header("Step 2: 名前フィルタ・JUSTDB用Excel出力")

names = df_current['権利部（甲区）氏名'].unique()
st.markdown("※現所有者全ての場合は「Select all」を選択ください")
selected_names = st.multiselect("フィルタする名前を選択してください", names)

exclude_cols = ['権利部（甲区）順位番号', '権利部（甲区）登記の目的', '権利部（甲区）種類']
output_cols = [c for c in df_current.columns if c not in exclude_cols]

if not selected_names:
    st.info("名前を選択すると、JUSTDB用Excelのダウンロードと現場データ生成に進めます。")
    st.stop()

filtered_df = df_current[df_current['権利部（甲区）氏名'].isin(selected_names)][output_cols]

st.write(f"**{len(filtered_df)}件** がマッチしました")
st.dataframe(filtered_df)

# JUSTDB用Excelダウンロード
if '所在' in filtered_df.columns:
    locations = filtered_df['所在'].unique()
    location_str = '_'.join(str(loc).replace(' ', '') for loc in locations[:3])
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        location_str = location_str.replace(ch, '')
    if len(locations) > 3:
        location_str += '_他'
    justdb_file_name = f"{location_str}_JUSTDB用.xlsx"
else:
    justdb_file_name = "filtered_data_JUSTDB用.xlsx"


def convert_df_to_excel(df_to_convert):
    output = BytesIO()
    df_to_convert.to_excel(output, index=False)
    return output.getvalue()


st.download_button(
    label=f"📥 JUSTDB用Excelをダウンロード ({justdb_file_name})",
    data=convert_df_to_excel(filtered_df),
    file_name=justdb_file_name,
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
)

# ── Step 3: 申請地/対面地/隣接地の分類 → 現場データ生成 ──
st.header("Step 3: 現場データ生成")
st.markdown("このまま現場データExcelが必要な方は使用ください")

if '地番' not in filtered_df.columns or '所在' not in filtered_df.columns:
    st.error("現場データ生成に必要な列（地番・所在）がありません。")
    st.stop()

地番一覧 = sorted(
    filtered_df['地番'].dropna().unique(),
    key=lambda x: int(''.join(filter(str.isdigit, str(x)))) if any(c.isdigit() for c in str(x)) else 0
)

col1, col2 = st.columns(2)
with col1:
    申請地リスト = st.multiselect("申請地を選択", 地番一覧)
with col2:
    残り = [x for x in 地番一覧 if x not in 申請地リスト]
    対面地リスト = st.multiselect("対面地を選択", 残り)

if st.button("現場データを出力"):
    if not 申請地リスト:
        st.warning("申請地を1つ以上選択してください。")
        st.stop()

    def classify(地番):
        if 地番 in 申請地リスト:
            return "申請地"
        elif 地番 in 対面地リスト:
            return "対面地"
        else:
            return "隣接地"

    df_work = filtered_df.copy()
    df_work['分類'] = df_work['地番'].apply(classify)
    df_work['sort_key'] = df_work['地番'].apply(
        lambda x: int(''.join(filter(str.isdigit, str(x)))) if pd.notnull(x) and any(c.isdigit() for c in str(x)) else 0
    )
    df_work['分類_order'] = pd.Categorical(df_work['分類'], ['申請地', '隣接地', '対面地'], ordered=True)
    df_work = df_work.sort_values(by=['分類_order', 'sort_key'])

    df_out = pd.DataFrame()
    df_out['B'] = df_work['分類']
    df_out['C'] = df_work['所在']
    df_out['D'] = df_work['地番']
    df_out['E'] = df_work['所在'].astype(str) + df_work['地番'].astype(str)
    df_out['F'] = df_work.get('地目', '')
    df_out['G'] = df_work.get('地積', '')
    df_out['H'] = df_work.get('権利部（甲区）住所', '')
    df_out['I'] = df_work.get('権利部（甲区）氏名', '')
    df_out['J'] = df_work.get('権利部（甲区）原因', '')

    wb = load_workbook('現場データ_テンプレート.xlsx')
    ws = wb['入力シート']

    row_start = 7
    for _, row in df_out.iterrows():
        for col_idx, val in enumerate(row, start=2):
            ws.cell(row=row_start, column=col_idx).value = val
        row_start += 1

    申請地_df = df_work[df_work['分類'] == '申請地']
    所在名 = str(申請地_df.iloc[0]['所在']).replace(' ', '').replace('/', '_') if not 申請地_df.empty else '未選択'
    genba_filename = f"現場データ({所在名}).xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    st.download_button(
        "📥 現場データをダウンロード",
        data=buffer.getvalue(),
        file_name=genba_filename,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
